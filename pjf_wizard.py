import os
import sys
import re
import getpass
from pathlib import Path
from collections import defaultdict

import openpyxl
from rich.console import Console
from rich.panel import Panel
from rich.prompt import Prompt
from rich.table import Table
from rich.progress import Progress, SpinnerColumn, BarColumn, TextColumn, TimeElapsedColumn, TimeRemainingColumn

import pjf_checker

console = Console()

FOLIO_RE = re.compile(r"^\s*(\d{1,6})\s*/\s*(\d{4})\s*$")
DATE_DDMMYYYY = re.compile(r"^\s*(\d{1,2})/(\d{1,2})/(\d{4})\s*$")


def normalize_path_input(raw: str) -> str:
    s = (raw or "").strip()
    if (s.startswith('"') and s.endswith('"')) or (s.startswith("'") and s.endswith("'")):
        s = s[1:-1]
    # macOS drag&drop escapes spaces with backslashes (Abner\ Dev)
    s = re.sub(r"\\([ ()\[\]{}&;\"'#$!])", r"\1", s)
    return s.strip()


def norm_txt(x: str) -> str:
    x = (x or "").strip().lower()
    x = x.replace("á", "a").replace("é", "e").replace("í", "i").replace("ó", "o").replace("ú", "u").replace("ü", "u")
    return " ".join(x.split())


def normalize_folio(v):
    if v is None:
        return None
    s = str(v).strip()
    m = FOLIO_RE.match(s)
    if not m:
        return None
    return f"{int(m.group(1))}/{m.group(2)}"


def excel_date_to_iso(v):
    if v is None:
        return None
    import datetime as dt
    if isinstance(v, dt.datetime):
        return v.date().isoformat()
    if isinstance(v, dt.date):
        return v.isoformat()

    s = str(v).strip()

    # Accept DD/MM/YY or DD/MM/YYYY
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{2,4})$", s)
    if m:
        dd = int(m.group(1))
        mm = int(m.group(2))
        yyyy = int(m.group(3))
        if yyyy < 100:
            yyyy = 2000 + yyyy
        return f"{yyyy:04d}-{mm:02d}-{dd:02d}"

    return None
    # Excel date -> python datetime/date
    import datetime as dt
    if isinstance(v, dt.datetime):
        return v.date().isoformat()
    if isinstance(v, dt.date):
        return v.isoformat()

    s = str(v).strip()
    m = DATE_DDMMYYYY.match(s)
    if m:
        dd = int(m.group(1))
        mm = int(m.group(2))
        yyyy = int(m.group(3))
        return f"{yyyy:04d}-{mm:02d}-{dd:02d}"
    return None


def detect_report_header(ws):
    # Find row containing "Número Exp" and "Tipo Asunto" (EGRESOS report style)
    for r in range(1, 80):
        row = [ws.cell(r, c).value for c in range(1, 20)]
        row_norm = [norm_txt(str(x)) if x is not None else "" for x in row]
        if "numero exp" in row_norm and "tipo asunto" in row_norm:
            col_map = {}
            for i, name in enumerate(row_norm, start=1):
                if name:
                    col_map[name] = i
            return r, col_map
    return None, None


def load_plan_from_excel(excel_path: Path):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active

    header_row, col_map = detect_report_header(ws)
    expected = {}
    plan = {}

    # REPORT MODE (EGRESOS)
    if header_row is not None:
        c_folio = col_map.get("numero exp")
        c_tipo = col_map.get("tipo asunto")
        c_fecha = col_map.get("valor fecha")

        by_tipo = defaultdict(list)
        r = header_row + 1
        while r <= ws.max_row:
            folio = normalize_folio(ws.cell(r, c_folio).value if c_folio else None)
            if folio is None:
                break

            tipo_raw = ws.cell(r, c_tipo).value if c_tipo else None
            tipo = str(tipo_raw).strip() if tipo_raw is not None else ""
            if tipo:
                by_tipo[tipo].append(folio)

                if c_fecha:
                    iso = excel_date_to_iso(ws.cell(r, c_fecha).value)
                    if iso:
                        expected[(tipo, folio)] = iso
            r += 1

        plan = dict(by_tipo)
        return plan, expected, "REPORT"

    # COLUMN MODE (original): row 1 = tipo headers, rows below = folios
    max_col = ws.max_column
    for c in range(1, max_col + 1):
        header = ws.cell(1, c).value
        if header is None:
            continue
        tipo = str(header).strip()
        if not tipo:
            continue
        folios = []
        r = 2
        while r <= ws.max_row:
            v = ws.cell(r, c).value
            if v is None:
                break
            f = normalize_folio(v)
            if f:
                folios.append(f)
            r += 1
        if folios:
            plan[tipo] = folios

    return plan, expected, "COLUMNS"


def open_path(path: Path):
    try:
        if sys.platform.startswith("darwin"):
            os.system(f'open "{path}"')
        elif os.name == "nt":
            os.startfile(str(path))  # type: ignore[attr-defined]
        else:
            os.system(f'xdg-open "{path}"')
    except Exception:
        pass


def main():
    console.print(
        Panel.fit(
            "[bold cyan]PJF Folio Checker[/bold cyan]\n"
            "[dim]Auto-detects report-style (EGRESOS) or column-style Excel[/dim]"
        )
    )

    console.print("\n[bold]Step 1 — Excel file[/bold]")
    console.print("[dim]Tip: drag & drop the file into this terminal and press Enter.[/dim]")
    excel_in = normalize_path_input(Prompt.ask("Excel file path"))
    excel_path = Path(excel_in).expanduser()
    if not excel_path.exists():
        console.print(f"[red]File not found:[/red] {excel_path}")
        raise SystemExit(1)

    console.print("\n[bold]Step 2 — Portal credentials[/bold]")
    user = Prompt.ask("Portal user").strip()
    password = getpass.getpass("Portal password: ")

    out_dir = excel_path.parent
    console.print(f"\n[bold]Outputs will be saved to:[/bold] [magenta]{out_dir}[/magenta]")

    plan, expected, mode = load_plan_from_excel(excel_path)
    if not plan:
        console.print("[red]No folios detected in Excel.[/red]")
        raise SystemExit(1)

    plan_tbl = Table(title="Plan (Detected from Excel)")
    plan_tbl.add_column("Tipo", style="cyan")
    plan_tbl.add_column("Folios", justify="right")
    total = 0
    for tipo in sorted(plan.keys()):
        n = len(plan[tipo])
        total += n
        plan_tbl.add_row(tipo, str(n))
    console.print(plan_tbl)

    console.print("\n[bold]Step 3 — Running (browser will open)[/bold]")

    counts = defaultdict(lambda: {})
    done = 0

    def cb(done_now, total_now, row):
        nonlocal done
        done = done_now
        st = row.get("status", "ERROR")
        tipo = row.get("tipo", "UNKNOWN")
        counts[tipo][st] = counts[tipo].get(st, 0) + 1
        prog.update(task_id, completed=done_now)

    with Progress(
        SpinnerColumn(),
        TextColumn("[bold]Checking[/bold] {task.completed}/{task.total}"),
        BarColumn(),
        TimeElapsedColumn(),
        TimeRemainingColumn(),
        console=console,
    ) as prog:
        task_id = prog.add_task("run", total=total)
        out_full, out_missing = pjf_checker.run_plan(
            plan=plan,
            expected_dates=expected,
            excel_label=excel_path.name,
            user=user,
            password=password,
            out_dir=str(out_dir),
            on_progress=cb,
        )

    console.print("\n[green]Done.[/green]")

    # Build a stable summary (show known statuses + any extras we saw)
    preferred = ["FOUND_MATCH", "FOUND_MISMATCH", "FOUND_NO_FECHA", "FOUND", "NOT_FOUND", "ERROR"]
    all_statuses = set()
    for tipo, d in counts.items():
        all_statuses.update(d.keys())
    extras = [x for x in sorted(all_statuses) if x not in preferred]
    columns = preferred + extras

    summary = Table(title="Summary by Tipo")
    summary.add_column("Tipo", style="cyan")
    for col in columns:
        summary.add_column(col, justify="right")

    for tipo in sorted(counts.keys()):
        c = counts[tipo]
        summary.add_row(tipo, *[str(c.get(col, 0)) for col in columns])

    console.print(summary)

    out_full_p = Path(out_full)
    out_missing_p = Path(out_missing)

    console.print(f"\n[bold]Results:[/bold]\n- {out_full_p}\n- {out_missing_p}")

    open_path(out_dir)
    open_path(out_full_p)
    open_path(out_missing_p)


if __name__ == "__main__":
    main()
