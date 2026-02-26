from pathlib import Path
import re

p = Path("pjf_checker.py")
s = p.read_text(encoding="utf-8")

start = s.find("def run_plan(")
if start == -1:
    raise SystemExit("Could not find def run_plan(...) in pjf_checker.py")

# Replace until the next top-level def (prefer def run(, else next def)
m = re.search(r"^def run\(", s[start:], flags=re.M)
end = start + m.start() if m else None
if not end:
    # fallback: next def after run_plan
    m2 = re.search(r"^def [a-zA-Z_]+\(", s[start+1:], flags=re.M)
    if not m2:
        raise SystemExit("Could not find end of run_plan block.")
    end = start + 1 + m2.start()

new_run_plan = r'''
def run_plan(*, plan, expected_dates, excel_label, user, password, out_dir, on_progress=None):
    """
    plan: {tipo: [folio,...]}
    expected_dates: {(tipo, folio): "YYYY-MM-DD"}  (from Excel Valor Fecha)
    Writes:
      - Folios_results.xlsx (includes FechaIngreso extracted + match status)
      - Folios_missing.xlsx (NOT_FOUND + MISMATCH + ERROR)
    Debug:
      - set env PJF_KEEP_TAB_SECONDS=3 to keep FOUND tab visible for 3 seconds
    """
    import os
    import time
    import re as _re
    from datetime import datetime
    from pathlib import Path as _Path
    from playwright.sync_api import sync_playwright

    keep_secs = 0.0
    try:
        keep_secs = float(os.environ.get("PJF_KEEP_TAB_SECONDS", "0") or "0")
    except Exception:
        keep_secs = 0.0

    def _js_click(locator):
        loc = locator.first
        loc.wait_for(state="attached", timeout=15000)
        loc.evaluate("el => el.click()")

    def _norm(x: str) -> str:
        x = (x or "").lower()
        x = x.replace("á","a").replace("é","e").replace("í","i").replace("ó","o").replace("ú","u").replace("ü","u")
        return " ".join(x.split())

    def _dismiss_not_found(page):
        # Click Aceptar on the visible Advertencia modal if present
        try:
            page.evaluate("""
() => {
  const visible = (el) => {
    if (!el) return false;
    const s = getComputedStyle(el);
    if (s.display === 'none' || s.visibility === 'hidden' || Number(s.opacity||'1') === 0) return false;
    const r = el.getBoundingClientRect();
    return r.width > 0 && r.height > 0;
  };
  const modals = Array.from(document.querySelectorAll('.modal'));
  const adv = modals.find(m => visible(m) && (m.innerText||'').includes('No existen datos para el expediente'));
  if (!adv) return;
  const btn = Array.from(adv.querySelectorAll('button,a')).find(b => (b.textContent||'').includes('Aceptar'));
  btn?.click();
  document.querySelectorAll('.modal-backdrop').forEach(e => e.remove());
  document.body.classList.remove('modal-open');
  document.body.style.removeProperty('padding-right');
}
""")
        except Exception:
            pass

    def _scroll_all(tab):
        try:
            tab.evaluate("""
() => {
  const els = Array.from(document.querySelectorAll('*'));
  for (const el of els) {
    const s = getComputedStyle(el);
    if (!s) continue;
    const oy = s.overflowY;
    if ((oy === 'auto' || oy === 'scroll') && el.scrollHeight > el.clientHeight + 20) {
      el.scrollTop = el.scrollHeight;
    }
  }
  window.scrollTo(0, document.body.scrollHeight);
}
""")
        except Exception:
            pass

    DATE = _re.compile(r"(\\d{2})/(\\d{2})/(\\d{4})")
    def _to_iso(text: str):
        m = DATE.search(text or "")
        if not m:
            return None
        dd, mm, yyyy = m.group(1), m.group(2), m.group(3)
        return f"{yyyy}-{mm}-{dd}"

    def _extract_fecha_ingreso(tab):
        # Try a few scroll passes (page has nested scroll areas)
        for _ in range(6):
            _scroll_all(tab)
            tab.wait_for_timeout(200)

        # Find a table that contains headers including "Fecha Ingreso" and "Archivo"
        fechas = []
        try:
            fechas = tab.evaluate(r"""
() => {
  const norm = (s) => (s||'').replace(/\\s+/g,' ').trim().toLowerCase();
  const tables = Array.from(document.querySelectorAll('table'));
  let target = null;

  for (const t of tables) {
    const ths = Array.from(t.querySelectorAll('th')).map(x => norm(x.textContent));
    if (ths.some(h => h.includes('fecha ingreso')) && ths.some(h => h.includes('archivo'))) {
      target = t; break;
    }
  }

  // Fallback: look near the heading "Listado de Resoluciones"
  if (!target) {
    const all = Array.from(document.querySelectorAll('*')).filter(el => el.children.length === 0);
    const h = all.find(el => norm(el.textContent).includes('listado de resoluciones'));
    if (h) {
      const t = h.closest('div')?.querySelector('table') || h.parentElement?.querySelector('table');
      if (t) target = t;
    }
  }

  if (!target) return [];

  const headers = Array.from(target.querySelectorAll('tr'))[0];
  if (!headers) return [];

  const cols = Array.from(headers.querySelectorAll('th,td')).map(x => norm(x.textContent));
  const idx = cols.findIndex(x => x.includes('fecha ingreso'));
  if (idx < 0) return [];

  const out = [];
  const rows = Array.from(target.querySelectorAll('tr')).slice(1);
  for (const r of rows) {
    const cells = Array.from(r.querySelectorAll('td,th'));
    if (cells.length <= idx) continue;
    const txt = (cells[idx].textContent||'').trim();
    out.push(txt);
  }
  return out;
}
""")
        except Exception:
            return []

        iso = []
        seen = set()
        for raw in fechas or []:
            v = _to_iso(raw)
            if v and v not in seen:
                seen.add(v)
                iso.append(v)
        return iso

    # Flatten plan into cases
    cases = []
    for tipo, folios in plan.items():
        for f in folios:
            cases.append((tipo, f, expected_dates.get((tipo, f))))

    total = len(cases)
    done = 0
    rows = []

    out_dir_p = _Path(out_dir).expanduser().resolve()
    out_full = out_dir_p / "Folios_results.xlsx"
    out_missing = out_dir_p / "Folios_missing.xlsx"

    with sync_playwright() as pwp:
        context = pwp.chromium.launch_persistent_context(
            user_data_dir="pw_profile_pjf_batch",
            headless=False,
            ignore_https_errors=True,
        )
        page = context.pages[0] if context.pages else context.new_page()

        # login
        auto_login_if_needed(page, user, password)

        # close modalidad if still present (we added this earlier)
        try:
            _ensure_modalidad_closed(page)
        except Exception:
            pass

        # open consulta (JS click to avoid overlay hangs)
        _js_click(page.get_by_text("Consulta de datos públicos de expedientes", exact=False))
        page.wait_for_load_state("domcontentloaded")

        # circuito + open filter modal
        select_circuito(page, CIRCUITO_TARGET)
        wait_for_filter_modal(page)
        modal = page.locator("div.modal:visible").filter(has_text=FILTER_TITLE).first

        # force organo (ddlOrgano=3457)
        _force_organo_octavo(modal, page)

        for (tipo, folio, expected_iso) in cases:
            done += 1
            status = "ERROR"
            notes = ""
            fecha_ingreso_iso = []
            match = None
            tab_url = ""

            try:
                # ensure filter modal is open
                if page.get_by_text(FILTER_TITLE, exact=False).count() == 0:
                    select_circuito(page, CIRCUITO_TARGET)
                    wait_for_filter_modal(page)
                    modal = page.locator("div.modal:visible").filter(has_text=FILTER_TITLE).first

                _force_organo_octavo(modal, page)

                # set tipo via Chosen (never select_option)
                try:
                    _set_tipo_via_chosen(modal, page, tipo)
                except Exception:
                    _set_tipo_via_chosen(modal, page, tipo)

                # fill folio
                inp = modal.get_by_placeholder("Ejemplo: 1/2026", exact=False)
                inp.wait_for(timeout=10000)
                inp.fill(folio)

                # click Buscar via JS (prevents hangs)
                before_pages = list(context.pages)
                before_count = len(before_pages)

                _js_click(modal.locator("button:has-text('Buscar'), a:has-text('Buscar')"))

                # wait for NOT_FOUND modal or new tab
                deadline = time.time() + 12.0
                new_tab = None

                while time.time() < deadline:
                    # NOT_FOUND check
                    if page.locator("text=No existen datos para el expediente").count() > 0:
                        _dismiss_not_found(page)
                        status = "NOT_FOUND"
                        break

                    pages_now = list(context.pages)
                    if len(pages_now) > before_count:
                        for p2 in pages_now[::-1]:
                            if p2 not in before_pages and p2 is not page:
                                new_tab = p2
                                break
                        if new_tab:
                            break

                    time.sleep(0.2)

                if new_tab:
                    try:
                        new_tab.wait_for_load_state("domcontentloaded", timeout=20000)
                    except Exception:
                        pass

                    tab_url = new_tab.url or ""
                    fecha_ingreso_iso = _extract_fecha_ingreso(new_tab)

                    if expected_iso and fecha_ingreso_iso:
                        match = expected_iso in set(fecha_ingreso_iso)
                        status = "FOUND_MATCH" if match else "FOUND_MISMATCH"
                    elif expected_iso and not fecha_ingreso_iso:
                        status = "FOUND_NO_FECHA"
                    else:
                        status = "FOUND"

                    if keep_secs > 0:
                        try:
                            new_tab.bring_to_front()
                        except Exception:
                            pass
                        time.sleep(keep_secs)

                    try:
                        new_tab.close()
                    except Exception:
                        pass

                if status == "ERROR":
                    notes = "No advertencia and no new tab within 12s."

            except Exception as e:
                status = "ERROR"
                notes = f"{type(e).__name__}: {e}"

            row = {
                "tipo": tipo,
                "folio": folio,
                "status": status,
                "expected_fecha_iso": expected_iso or "",
                "fecha_ingreso_iso": ", ".join(fecha_ingreso_iso) if fecha_ingreso_iso else "",
                "match": match if match is not None else "",
                "tab_url": tab_url,
                "checked_at": datetime.now().isoformat(timespec="seconds"),
                "notes": notes,
            }
            rows.append(row)

            if on_progress:
                on_progress(done, total, row)

            # checkpoint every 25 so crashes don't lose everything
            if done % 25 == 0:
                _write_results(out_full, out_missing, rows)

        try:
            context.close()
        except Exception:
            pass

    _write_results(out_full, out_missing, rows)
    return str(out_full), str(out_missing)


def _write_results(out_full, out_missing, rows):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    ws.append(["Tipo", "Folio", "Status", "Expected_ValorFecha_ISO", "FechaIngreso_ISO_List", "Match", "TabURL", "CheckedAt", "Notes"])
    for r in rows:
        ws.append([r["tipo"], r["folio"], r["status"], r["expected_fecha_iso"], r["fecha_ingreso_iso"], r["match"], r["tab_url"], r["checked_at"], r["notes"]])
        ws.cell(row=ws.max_row, column=2).number_format = "@"
    wb.save(out_full)

    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "FollowUp"
    ws2.append(["Tipo", "Folio", "Status", "Expected_ValorFecha_ISO", "FechaIngreso_ISO_List"])
    for r in rows:
        if r["status"] in ("NOT_FOUND", "FOUND_MISMATCH", "FOUND_NO_FECHA", "ERROR"):
            ws2.append([r["tipo"], r["folio"], r["status"], r["expected_fecha_iso"], r["fecha_ingreso_iso"]])
            ws2.cell(row=ws2.max_row, column=2).number_format = "@"
    wb2.save(out_missing)
'''

s2 = s[:start] + new_run_plan + s[end:]
p.write_text(s2, encoding="utf-8")
print("Replaced run_plan() with Fecha Ingreso extraction + non-hanging Buscar + checkpoints.")
