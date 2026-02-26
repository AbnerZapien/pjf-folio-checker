
import re


# --- Safe defaults (so run_plan won't crash if constants were removed by a restore/patch) ---
try:
    CIRCUITO_TARGET
except NameError:
    CIRCUITO_TARGET = "SEXTO CIRCUITO"

try:
    FILTER_TITLE
except NameError:
    FILTER_TITLE = "Filtro Expediente"

try:
    NOT_FOUND_TEXT
except NameError:
    NOT_FOUND_TEXT = "No existen datos para el expediente"
# --- End defaults ---


def select_circuito(page, circuito_text: str):
    # Click the circuito control next to label "Circuito:" and select by typing + Enter
    label = page.locator("xpath=//*[normalize-space()='Circuito:']")
    label.wait_for(timeout=20000)
    ctrl = label.locator("xpath=following::*[self::div or self::span or self::input][1]")
    ctrl.click(force=True)
    page.keyboard.type(circuito_text, delay=25)
    page.keyboard.press("Enter")

def wait_for_filter_modal(page):
    page.get_by_text(FILTER_TITLE, exact=False).wait_for(timeout=20000)

BASE_URL = "https://www.serviciosenlinea.pjf.gob.mx/"
LOGIN_MENU_TEXT = "Ingresa al Portal"
LOGIN_TARGET_TEXT = "Juzgados de Distrito y Tribunales de Circuito"

def _setup_auto_dialog_accept(page):
    def _on_dialog(d):
        try:
            d.accept()
        except Exception:
            pass
    page.on("dialog", _on_dialog)

def _is_logged_in(page) -> bool:
    return page.get_by_text("Consulta de datos públicos de expedientes", exact=False).count() > 0

def _cleanup_backdrops(page):
    try:
        page.evaluate("""
          document.querySelectorAll('.modal-backdrop').forEach(e => e.remove());
          document.body.classList.remove('modal-open');
          document.body.style.removeProperty('padding-right');
        """)
    except Exception:
        pass


def _js_click(locator):
    loc = locator.first
    loc.wait_for(state="attached", timeout=15000)
    loc.evaluate("el => el.click()")

def _ensure_modalidad_closed(page):
    for _ in range(10):
        try:
            visible = page.evaluate("""
() => {
  const m = document.getElementById('modalModalidad');
  if (!m) return false;
  const st = window.getComputedStyle(m);
  const shown = st && st.display !== 'none' && st.visibility !== 'hidden' && st.opacity !== '0';
  const cls = m.classList.contains('in') || m.classList.contains('show');
  const disp = (m.style && (m.style.display === 'block'));
  return shown && (cls || disp);
}
""")
            if not visible:
                try:
                    page.evaluate("""
() => {
  const m = document.getElementById('modalModalidad');
  if (m) { m.classList.remove('in','show'); m.style.display='none'; m.setAttribute('aria-hidden','true'); }
  document.querySelectorAll('.modal-backdrop').forEach(e => e.remove());
  document.body.classList.remove('modal-open');
  document.body.style.removeProperty('padding-right');
}
""")
                except Exception:
                    pass
                return

            try:
                page.keyboard.press("Escape")
            except Exception:
                pass

            try:
                page.evaluate("""
() => {
  const m = document.getElementById('modalModalidad');
  if (m) { m.classList.remove('in','show'); m.style.display='none'; m.setAttribute('aria-hidden','true'); }
  document.querySelectorAll('.modal-backdrop').forEach(e => e.remove());
  document.body.classList.remove('modal-open');
  document.body.style.removeProperty('padding-right');
}
""")
            except Exception:
                pass

            try:
                _cleanup_backdrops(page)
            except Exception:
                pass

            page.wait_for_timeout(200)
        except Exception:
            try:
                page.evaluate("""
() => {
  const m = document.getElementById('modalModalidad');
  if (m) { m.classList.remove('in','show'); m.style.display='none'; m.setAttribute('aria-hidden','true'); }
  document.querySelectorAll('.modal-backdrop').forEach(e => e.remove());
  document.body.classList.remove('modal-open');
  document.body.style.removeProperty('padding-right');
}
""")
            except Exception:
                pass
            try:
                _cleanup_backdrops(page)
            except Exception:
                pass
            page.wait_for_timeout(200)

    try:
        _cleanup_backdrops(page)
    except Exception:
        pass


def auto_login_if_needed(page, user: str, password: str):
    _setup_auto_dialog_accept(page)

    page.goto(BASE_URL, wait_until="domcontentloaded")

    if _is_logged_in(page):
        return

    # Open the dropdown
    page.get_by_text(re.compile(r"Ingresa al Portal|Ingresar al Portal", re.I), exact=False).click()

    # Click the dropdown link (use role=link to avoid strict-mode collisions with description text)
    page.get_by_role("link", name=re.compile(r"Juzgados de Distrito", re.I)).click()

    # Login form
    u = page.locator("#UserName")
    pw = page.locator("#UserPassword")
    u.wait_for(state="visible", timeout=20000)
    pw.wait_for(state="visible", timeout=20000)

    u.fill(user)
    pw.fill(password)

    page.get_by_role("button", name=re.compile(r"^Ingresar$", re.I)).click()

    # Profile modal: click Persona Física (sometimes appears a second later)
    try:
        page.get_by_text("Seleccione un perfil", exact=False).wait_for(timeout=20000)

        # Prefer clicking the tile container; fallback to any text match
        try:
            page.locator("span.btnDemandaA", has_text=re.compile(r"Persona\s+F[ií]sica", re.I)).first.click(force=True, timeout=15000)
        except Exception:
            page.get_by_text(re.compile(r"Persona\s+F[ií]sica", re.I), exact=False).first.click(force=True, timeout=15000)

        # Alert after selecting profile is auto-accepted by dialog handler

        # Wait for modal to disappear (best effort, then force-clean overlays)
        try:
            page.locator("#modalModalidad").wait_for(state="hidden", timeout=12000)
        except Exception:
            pass
        _cleanup_backdrops(page)

    except Exception:
        # Sometimes it skips profile selection and goes straight in
        _cleanup_backdrops(page)

    # Confirm we are in
    page.get_by_text("Consulta de datos públicos de expedientes", exact=False).wait_for(timeout=30000)

import re
import time
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook, Workbook
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout

ORGANO_TARGET_VALUE = "3457"  # Juzgado Octavo...

BASE_URL = "https://www.serviciosenlinea.pjf.gob.mx/"
LOGIN_MENU_REGEX = re.compile(r"Ingres(a|ar)\s+al\s+Portal", re.I)
LOGIN_TARGET_TEXT = "Juzgados de Distrito y Tribunales de Circuito"

FILTER_TITLE = "Filtro Expediente"
NOT_FOUND_TEXT = "No existen datos para el expediente"

EXP_RE = re.compile(r"^\s*(\d{1,6})\s*/\s*(\d{4})\s*$")

def _normalize_folio(v):
    if v is None:
        return None
    s = str(v).strip()
    m = EXP_RE.match(s)
    if not m:
        return None
    return f"{int(m.group(1))}/{m.group(2)}"

def read_folios_by_tipo_xlsx(excel_path: str) -> dict[str, list[str]]:
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active

    tipos: dict[str, list[str]] = {}
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header is None:
            continue
        tipo = str(header).strip()
        if not tipo:
            continue

        folios: list[str] = []
        row = 2
        while True:
            v = ws.cell(row=row, column=col).value
            if v is None:
                break
            nf = _normalize_folio(v)
            if nf:
                folios.append(nf)
            row += 1

        seen, out = set(), []
        for f in folios:
            if f not in seen:
                seen.add(f)
                out.append(f)

        if out:
            tipos[tipo] = out

    return tipos

def _write_outputs(rows, out_dir: Path):
    out_full = out_dir / "Folios_results.xlsx"
    out_missing = out_dir / "Folios_missing.xlsx"

    wb1 = Workbook()
    ws1 = wb1.active
    ws1.title = "Results"
    ws1.append(["Tipo", "Folio", "Found", "Status", "CheckedAt", "Notes"])
    for r in rows:
        ws1.append([r["tipo"], r["folio"], r["found"], r["status"], r["checked_at"], r["notes"] or ""])
    wb1.save(out_full)

    missing = [r for r in rows if (not r["found"] and r["status"] in ("NOT_FOUND", "ERROR"))]
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "MissingOrError"
    ws2.append(["Tipo", "Folio", "Status"])
    for r in missing:
        ws2.append([r["tipo"], r["folio"], r["status"]])
        ws2.cell(row=ws2.max_row, column=2).number_format = "@"
    wb2.save(out_missing)

    return out_full, out_missing

def _cleanup_overlays(page):
    try:
        page.evaluate("""
          document.querySelectorAll('.modal-backdrop').forEach(e => e.remove());
          document.body.classList.remove('modal-open');
          document.body.style.removeProperty('padding-right');
        """)
    except Exception:
        pass

def _close_extra_pages(context, main_page):
    for p in list(context.pages):
        if p is main_page:
            continue
        try:
            p.close()
        except Exception:
            pass

def _setup_auto_dialog_accept(page):
    def _on_dialog(d):
        try:
            d.accept()
        except Exception:
            pass
    page.on("dialog", _on_dialog)

def _is_logged_in(page) -> bool:
    return page.get_by_text("Consulta de datos públicos de expedientes", exact=False).count() > 0

def _click_login_entry(page):
    locs = [
        page.get_by_text(LOGIN_MENU_REGEX, exact=False),
        page.get_by_role("link", name=re.compile("Ingres", re.I)),
        page.get_by_role("button", name=re.compile("Ingres", re.I)),
        page.locator("a:has-text('Ingres')"),
        page.locator("button:has-text('Ingres')"),
    ]
    for loc in locs:
        try:
            if loc.count() > 0:
                loc.first.click(force=True, timeout=8000)
                return
        except Exception:
            pass
    raise RuntimeError("Could not click login entry (Ingresa al Portal).")

def _ensure_profile_selected(page):
    # Optimized but stable timeouts (your latest values)
    modal = page.locator("#modalModalidad")
    try:
        modal.wait_for(state="visible", timeout=4000)
    except PWTimeout:
        return

    for _ in range(6):
        try:
            page.evaluate("""
              const m = document.getElementById('modalModalidad');
              if (!m) return;
              const norm = s => (s||'').replace(/\\s+/g,' ').trim().toLowerCase();
              const leafs = Array.from(m.querySelectorAll('*')).filter(el => el.children.length === 0);
              const node = leafs.find(el => {
                const t = norm(el.textContent);
                return t === 'persona física' || t === 'persona fisica';
              });
              if (node) {
                const clickable = node.closest('[onclick]') || node.closest('a,button') || node.closest('div') || node;
                clickable.click();
              }
            """)
        except Exception:
            pass

        try:
            modal.wait_for(state="hidden", timeout=1200)
            _cleanup_overlays(page)
            return
        except Exception:
            _cleanup_overlays(page)
            page.wait_for_timeout(250)

    # last resort: hide to unblock
    try:
        page.evaluate("""
          const m = document.getElementById("modalModalidad");
          if (m) { m.style.display="none"; m.classList.remove("in","show"); }
        """)
    except Exception:
        pass
    _cleanup_overlays(page)

def _auto_login_if_needed(page, user: str, password: str):
    page.goto(BASE_URL, wait_until="domcontentloaded")
    page.wait_for_timeout(300)

    if _is_logged_in(page):
        return

    _click_login_entry(page)
    page.wait_for_timeout(250)

    page.get_by_role("link", name=re.compile(LOGIN_TARGET_TEXT, re.I)).click(timeout=15000)
    page.wait_for_load_state("domcontentloaded")

    u = page.locator("#UserName")
    p = page.locator("#UserPassword")
    u.wait_for(state="visible", timeout=20000)
    p.wait_for(state="visible", timeout=20000)

    u.fill(user)
    p.fill(password)
    page.get_by_role("button", name=re.compile(r"^Ingresar$", re.I)).click()

    _ensure_profile_selected(page)

    marker = page.get_by_text("Consulta de datos públicos de expedientes", exact=False)
    try:
        marker.wait_for(timeout=8000)
    except PWTimeout:
        marker.wait_for(timeout=30000)

def _click_menu_consulta(page):
    _cleanup_overlays(page)
    page.get_by_text("Consulta de datos públicos de expedientes", exact=False).first.evaluate("el => el.click()")

def _select_circuito(page, circuito_text):
    label = page.locator("xpath=//*[normalize-space()='Circuito:']")
    label.wait_for(timeout=12000)
    ctrl = label.locator("xpath=following::*[self::div or self::span or self::input][1]")
    ctrl.click(force=True)
    page.keyboard.type(circuito_text, delay=25)
    page.keyboard.press("Enter")

def _wait_for_filter_modal(page):
    page.get_by_text(FILTER_TITLE, exact=False).wait_for(timeout=12000)

def _modal_root(page):
    m = page.locator("xpath=//*[contains(normalize-space(),'Filtro Expediente')]/ancestor::*[contains(@class,'modal')][1]")
    if m.count() == 0:
        m = page.locator("xpath=//*[contains(normalize-space(),'Filtro Expediente')]/ancestor::*[1]")
    return m.first

def _set_tipo_chosen(modal, page, tipo_text):
    chosen = modal.locator("#ddlTipoAsunto_chosen")
    if chosen.count() == 0:
        chosen = modal.locator("select#ddlTipoAsunto").locator("xpath=following-sibling::div[contains(@class,'chosen-container')][1]")
    chosen.first.wait_for(timeout=10000)
    chosen.first.click(force=True)
    search = chosen.first.locator("input[type='text']")
    if search.count() > 0:
        search.first.fill(tipo_text)
    else:
        page.keyboard.type(tipo_text, delay=25)
    page.keyboard.press("Enter")

def _fill_folio(modal, folio):
    inp = modal.get_by_placeholder("Ejemplo: 1/2026", exact=False)
    inp.wait_for(timeout=10000)
    inp.fill(folio)

def _click_buscar(modal):
    btn = modal.locator("button:has-text('Buscar'), a:has-text('Buscar')").first
    btn.wait_for(state="attached", timeout=10000)
    btn.evaluate("el => el.click()")

def _advertencia_visible_js(page) -> bool:
    try:
        return bool(page.evaluate("""
          (txt) => {
            const visible = (el) => {
              if (!el) return false;
              const s = getComputedStyle(el);
              if (!s) return false;
              if (s.display === 'none' || s.visibility === 'hidden' || Number(s.opacity||'1') === 0) return false;
              const r = el.getBoundingClientRect();
              return r.width > 0 && r.height > 0;
            };
            const m = Array.from(document.querySelectorAll('.modal'))
              .find(x => visible(x) && (x.innerText||'').includes(txt));
            return !!m;
          }
        """, NOT_FOUND_TEXT))
    except Exception:
        return False

def _dismiss_advertencia_js(page):
    if not _advertencia_visible_js(page):
        return False
    try:
        page.evaluate("""
          (txt) => {
            const visible = (el) => {
              if (!el) return false;
              const s = getComputedStyle(el);
              if (!s) return false;
              if (s.display === 'none' || s.visibility === 'hidden' || Number(s.opacity||'1') === 0) return false;
              const r = el.getBoundingClientRect();
              return r.width > 0 && r.height > 0;
            };
            const m = Array.from(document.querySelectorAll('.modal'))
              .find(x => visible(x) && (x.innerText||'').includes(txt));
            if (!m) return;
            const btn = Array.from(m.querySelectorAll('button,a'))
              .find(b => (b.textContent||'').includes('Aceptar'));
            btn?.click();
            document.querySelectorAll('.modal-backdrop').forEach(e => e.remove());
            document.body.classList.remove('modal-open');
            document.body.style.removeProperty('padding-right');
          }
        """, NOT_FOUND_TEXT)
    except Exception:
        pass
    page.wait_for_timeout(200)
    _cleanup_overlays(page)
    return True


def _ensure_organo_octavo(modal, page):
    sel = modal.locator("#ddlOrgano").first
    if sel.count() == 0:
        try:
            modal.screenshot(path="debug_organo_failed.png")
        except Exception:
            pass
        raise RuntimeError("ddlOrgano not found inside modal (debug_organo_failed.png).")

    sel.wait_for(state="attached", timeout=8000)
    sel.evaluate("(el, v) => { el.value = v; el.dispatchEvent(new Event('change', {bubbles:true})); }", ORGANO_TARGET_VALUE)
    page.wait_for_timeout(250)

    # Update Chosen UI (best effort)
    try:
        page.evaluate("() => { const el=document.querySelector('#ddlOrgano'); try { if (window.jQuery && el) window.jQuery(el).trigger('chosen:updated'); } catch(e) {} }")
    except Exception:
        pass

    selected = sel.evaluate("el => (el.selectedOptions && el.selectedOptions[0] ? (el.selectedOptions[0].textContent||'').trim() : '')")

    if ("Octavo" not in selected) or ("Puebla" not in selected):
        Path("debug_organo_selected.txt").write_text("Selected\n" + selected, encoding="utf-8")
        try:
            modal.screenshot(path="debug_organo_failed.png")
        except Exception:
            pass
        raise RuntimeError("Órgano did not stick to Juzgado Octavo (debug_organo_selected.txt + debug_organo_failed.png).")



def _force_organo_octavo(modal, page):
    # ddlOrgano is hidden by Chosen, so set via JS
    sel = modal.locator("#ddlOrgano").first
    if sel.count() == 0:
        return False
    sel.evaluate("(el, v) => { el.value=v; el.dispatchEvent(new Event('change', {bubbles:true})); }", ORGANO_TARGET_VALUE)
    page.wait_for_timeout(250)
    try:
        page.evaluate("() => { const el=document.querySelector('#ddlOrgano'); try { if (window.jQuery && el) window.jQuery(el).trigger('chosen:updated'); } catch(e) {} }")
    except Exception:
        pass
    return True



def _set_tipo_via_chosen(modal, page, tipo_label: str):
    # Works even when the real <select> is hidden by Chosen
    chosen = modal.locator("#ddlTipoAsunto_chosen").first
    if chosen.count() == 0:
        # fallback: chosen container after select
        chosen = modal.locator("select#ddlTipoAsunto").locator("xpath=following-sibling::div[contains(@class,'chosen-container')][1]").first
    chosen.wait_for(state="attached", timeout=10000)
    chosen.click(force=True)

    search = chosen.locator("input[type='text']").first
    try:
        search.wait_for(state="attached", timeout=5000)
        search.fill(tipo_label)
    except Exception:
        page.keyboard.type(tipo_label, delay=25)

    page.keyboard.press("Enter")
    page.wait_for_timeout(200)



def run_plan(*, plan, expected_dates, excel_label, user, password, out_dir, on_progress=None):
    """
    plan: {tipo: [folio,...]}
    expected_dates: {(tipo, folio): "YYYY-MM-DD"}  (from Excel Valor Fecha)
    Writes:
      - Folios_results.xlsx (includes FechaIngreso extracted + match status)
      - Folios_missing.xlsx (NOT_FOUND + MISMATCH + ERROR)
    Debug:
      - set env PJF_KEEP_TAB_SECONDS=3 to keep FOUND tab visible for 3 seconds
    """
    import os
    import time
    import re as _re
    from datetime import datetime
    from pathlib import Path as _Path
    from playwright.sync_api import sync_playwright

    keep_secs = 0.0
    try:
        keep_secs = float(os.environ.get("PJF_KEEP_TAB_SECONDS", "0") or "0")
    except Exception:
        keep_secs = 0.0

    def _js_click(locator):
        loc = locator.first
        loc.wait_for(state="attached", timeout=15000)
        loc.evaluate("el => el.click()")

    def _norm(x: str) -> str:
        x = (x or "").lower()
        x = x.replace("á","a").replace("é","e").replace("í","i").replace("ó","o").replace("ú","u").replace("ü","u")
        return " ".join(x.split())

    def _dismiss_not_found(page):
        # Click Aceptar on the visible Advertencia modal if present
        try:
            page.evaluate("""
() => {
  const visible = (el) => {
    if (!el) return false;
    const s = getComputedStyle(el);
    if (s.display === 'none' || s.visibility === 'hidden' || Number(s.opacity||'1') === 0) return false;
    const r = el.getBoundingClientRect();
    return r.width > 0 && r.height > 0;
  };
  const modals = Array.from(document.querySelectorAll('.modal'));
  const adv = modals.find(m => visible(m) && (m.innerText||'').includes('No existen datos para el expediente'));
  if (!adv) return;
  const btn = Array.from(adv.querySelectorAll('button,a')).find(b => (b.textContent||'').includes('Aceptar'));
  btn?.click();
  document.querySelectorAll('.modal-backdrop').forEach(e => e.remove());
  document.body.classList.remove('modal-open');
  document.body.style.removeProperty('padding-right');
}
""")
        except Exception:
            pass

    def _scroll_all(tab):
        try:
            tab.evaluate("""
() => {
  const els = Array.from(document.querySelectorAll('*'));
  for (const el of els) {
    const s = getComputedStyle(el);
    if (!s) continue;
    const oy = s.overflowY;
    if ((oy === 'auto' || oy === 'scroll') && el.scrollHeight > el.clientHeight + 20) {
      el.scrollTop = el.scrollHeight;
    }
  }
  window.scrollTo(0, document.body.scrollHeight);
}
""")
        except Exception:
            pass

    DATE = _re.compile(r"(\\d{2})/(\\d{2})/(\\d{4})")
    def _to_iso(text: str):
        m = DATE.search(text or "")
        if not m:
            return None
        dd, mm, yyyy = m.group(1), m.group(2), m.group(3)
        return f"{yyyy}-{mm}-{dd}"

    def _extract_fecha_ingreso(tab):
        # Try a few scroll passes (page has nested scroll areas)
        for _ in range(6):
            _scroll_all(tab)
            tab.wait_for_timeout(200)

        # Find a table that contains headers including "Fecha Ingreso" and "Archivo"
        fechas = []
        try:
            fechas = tab.evaluate(r"""
() => {
  const norm = (s) => (s||'').replace(/\\s+/g,' ').trim().toLowerCase();
  const tables = Array.from(document.querySelectorAll('table'));
  let target = null;

  for (const t of tables) {
    const ths = Array.from(t.querySelectorAll('th')).map(x => norm(x.textContent));
    if (ths.some(h => h.includes('fecha ingreso')) && ths.some(h => h.includes('archivo'))) {
      target = t; break;
    }
  }

  // Fallback: look near the heading "Listado de Resoluciones"
  if (!target) {
    const all = Array.from(document.querySelectorAll('*')).filter(el => el.children.length === 0);
    const h = all.find(el => norm(el.textContent).includes('listado de resoluciones'));
    if (h) {
      const t = h.closest('div')?.querySelector('table') || h.parentElement?.querySelector('table');
      if (t) target = t;
    }
  }

  if (!target) return [];

  const headers = Array.from(target.querySelectorAll('tr'))[0];
  if (!headers) return [];

  const cols = Array.from(headers.querySelectorAll('th,td')).map(x => norm(x.textContent));
  const idx = cols.findIndex(x => x.includes('fecha ingreso'));
  if (idx < 0) return [];

  const out = [];
  const rows = Array.from(target.querySelectorAll('tr')).slice(1);
  for (const r of rows) {
    const cells = Array.from(r.querySelectorAll('td,th'));
    if (cells.length <= idx) continue;
    const txt = (cells[idx].textContent||'').trim();
    out.push(txt);
  }
  return out;
}
""")
        except Exception:
            return []

        iso = []
        seen = set()
        for raw in fechas or []:
            v = _to_iso(raw)
            if v and v not in seen:
                seen.add(v)
                iso.append(v)
        return iso

    # Flatten plan into cases
    cases = []
    for tipo, folios in plan.items():
        for f in folios:
            cases.append((tipo, f, expected_dates.get((tipo, f))))

    total = len(cases)
    done = 0
    rows = []

    out_dir_p = _Path(out_dir).expanduser().resolve()
    out_full = out_dir_p / "Folios_results.xlsx"
    out_missing = out_dir_p / "Folios_missing.xlsx"

    with sync_playwright() as pwp:
        context = pwp.chromium.launch_persistent_context(
            user_data_dir="pw_profile_pjf_batch",
            headless=False,
            ignore_https_errors=True,
        )
        page = context.pages[0] if context.pages else context.new_page()

        # login
        auto_login_if_needed(page, user, password)

        # close modalidad if still present (we added this earlier)
        try:
            _ensure_modalidad_closed(page)
        except Exception:
            pass

        # open consulta (JS click to avoid overlay hangs)
        _js_click(page.get_by_text("Consulta de datos públicos de expedientes", exact=False))
        page.wait_for_load_state("domcontentloaded")

        # circuito + open filter modal
        select_circuito(page, CIRCUITO_TARGET)
        wait_for_filter_modal(page)
        modal = page.locator("div.modal:visible").filter(has_text=FILTER_TITLE).first

        # force organo (ddlOrgano=3457)
        _force_organo_octavo(modal, page)

        for (tipo, folio, expected_iso) in cases:
            done += 1
            status = "ERROR"
            notes = ""
            fecha_ingreso_iso = []
            match = None
            tab_url = ""

            try:
                # ensure filter modal is open
                if page.get_by_text(FILTER_TITLE, exact=False).count() == 0:
                    select_circuito(page, CIRCUITO_TARGET)
                    wait_for_filter_modal(page)
                    modal = page.locator("div.modal:visible").filter(has_text=FILTER_TITLE).first

                _force_organo_octavo(modal, page)

                # set tipo via Chosen (never select_option)
                try:
                    _set_tipo_via_chosen(modal, page, tipo)
                except Exception:
                    _set_tipo_via_chosen(modal, page, tipo)

                # fill folio
                inp = modal.get_by_placeholder("Ejemplo: 1/2026", exact=False)
                inp.wait_for(timeout=10000)
                inp.fill(folio)

                # click Buscar via JS (prevents hangs)
                before_pages = list(context.pages)
                before_count = len(before_pages)

                _js_click(modal.locator("button:has-text('Buscar'), a:has-text('Buscar')"))

                # wait for NOT_FOUND modal or new tab
                deadline = time.time() + 12.0
                new_tab = None

                while time.time() < deadline:
                    # NOT_FOUND check
                    if page.locator("text=No existen datos para el expediente").count() > 0:
                        _dismiss_not_found(page)
                        status = "NOT_FOUND"
                        break

                    pages_now = list(context.pages)
                    if len(pages_now) > before_count:
                        for p2 in pages_now[::-1]:
                            if p2 not in before_pages and p2 is not page:
                                new_tab = p2
                                break
                        if new_tab:
                            break

                    time.sleep(0.2)

                if new_tab:
                    try:
                        new_tab.wait_for_load_state("domcontentloaded", timeout=20000)
                    except Exception:
                        pass

                    tab_url = new_tab.url or ""
                    fecha_ingreso_iso = _extract_fecha_ingreso(new_tab)

                    if expected_iso and fecha_ingreso_iso:
                        match = expected_iso in set(fecha_ingreso_iso)
                        status = "FOUND_MATCH" if match else "FOUND_MISMATCH"
                    elif expected_iso and not fecha_ingreso_iso:
                        status = "FOUND_NO_FECHA"
                    else:
                        status = "FOUND"

                    if keep_secs > 0:
                        try:
                            new_tab.bring_to_front()
                        except Exception:
                            pass
                        time.sleep(keep_secs)

                    try:
                        new_tab.close()
                    except Exception:
                        pass

                if status == "ERROR":
                    notes = "No advertencia and no new tab within 12s."

            except Exception as e:
                status = "ERROR"
                notes = f"{type(e).__name__}: {e}"

            row = {
                "tipo": tipo,
                "folio": folio,
                "status": status,
                "expected_fecha_iso": expected_iso or "",
                "fecha_ingreso_iso": ", ".join(fecha_ingreso_iso) if fecha_ingreso_iso else "",
                "match": match if match is not None else "",
                "tab_url": tab_url,
                "checked_at": datetime.now().isoformat(timespec="seconds"),
                "notes": notes,
            }
            rows.append(row)

            if on_progress:
                on_progress(done, total, row)

            # checkpoint every 25 so crashes don't lose everything
            if done % 25 == 0:
                _write_results(out_full, out_missing, rows)

        try:
            context.close()
        except Exception:
            pass

    _write_results(out_full, out_missing, rows)
    return str(out_full), str(out_missing)


def _write_results(out_full, out_missing, rows):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    ws.append(["Tipo", "Folio", "Status", "Expected_ValorFecha_ISO", "FechaIngreso_ISO_List", "Match", "TabURL", "CheckedAt", "Notes"])
    for r in rows:
        ws.append([r["tipo"], r["folio"], r["status"], r["expected_fecha_iso"], r["fecha_ingreso_iso"], r["match"], r["tab_url"], r["checked_at"], r["notes"]])
        ws.cell(row=ws.max_row, column=2).number_format = "@"
    wb.save(out_full)

    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "FollowUp"
    ws2.append(["Tipo", "Folio", "Status", "Expected_ValorFecha_ISO", "FechaIngreso_ISO_List"])
    for r in rows:
        if r["status"] in ("NOT_FOUND", "FOUND_MISMATCH", "FOUND_NO_FECHA", "ERROR"):
            ws2.append([r["tipo"], r["folio"], r["status"], r["expected_fecha_iso"], r["fecha_ingreso_iso"]])
            ws2.cell(row=ws2.max_row, column=2).number_format = "@"
    wb2.save(out_missing)
def run(excel_path: str, user: str, password: str, circuito: str = "SEXTO CIRCUITO",
        profile_dir: str = "pw_profile_pjf_multi", on_progress=None):
    tipos = read_folios_by_tipo_xlsx(excel_path)
    if not tipos:
        raise RuntimeError("No tipos/folios found. Headers must be in row 1; folios start row 2.")

    out_dir = Path(excel_path).expanduser().resolve().parent
    rows = []

    total = sum(len(v) for v in tipos.values())
    done = 0

    with sync_playwright() as p:
        context = p.chromium.launch_persistent_context(
            user_data_dir=str((Path.cwd() / profile_dir).resolve()),
            headless=False,
            ignore_https_errors=True,
        )
        page = context.pages[0] if context.pages else context.new_page()
        _setup_auto_dialog_accept(page)

        _auto_login_if_needed(page, user, password)
        _ensure_profile_selected(page)

        _click_menu_consulta(page)
        page.wait_for_timeout(300)

        _select_circuito(page, circuito)
        _wait_for_filter_modal(page)
        modal = _modal_root(page)

        try:
            for tipo, folios in tipos.items():
                _set_tipo_chosen(modal, page, tipo)

                for folio in folios:
                    done += 1
                    _close_extra_pages(context, page)
                    _dismiss_advertencia_js(page)

                    if page.get_by_text(FILTER_TITLE, exact=False).count() == 0:
                        _select_circuito(page, circuito)
                        _wait_for_filter_modal(page)
                        modal = _modal_root(page)
                        _set_tipo_chosen(modal, page, tipo)

                    found = False
                    status = "ERROR"
                    notes = ""

                    _fill_folio(modal, folio)

                    before_pages = list(context.pages)
                    before_count = len(before_pages)

                    _click_buscar(modal)

                    deadline = time.time() + 8.0
                    new_tab = None

                    while time.time() < deadline:
                        if _advertencia_visible_js(page):
                            _dismiss_advertencia_js(page)
                            found = False
                            status = "NOT_FOUND"
                            break

                        pages_now = list(context.pages)
                        if len(pages_now) > before_count:
                            for p2 in pages_now[::-1]:
                                if p2 not in before_pages and p2 is not page:
                                    new_tab = p2
                                    break
                            if new_tab:
                                break

                        time.sleep(0.15)

                    if new_tab:
                        try:
                            new_tab.wait_for_load_state("domcontentloaded", timeout=15000)
                        except Exception:
                            pass
                        try:
                            new_tab.close()
                        except Exception:
                            pass
                        _close_extra_pages(context, page)
                        found = True
                        status = "FOUND"

                    if status == "ERROR":
                        notes = "No visible Advertencia and no new tab within 8s."

                    row = {
                        "tipo": tipo,
                        "folio": folio,
                        "found": found,
                        "status": status,
                        "checked_at": datetime.now().isoformat(timespec="seconds"),
                        "notes": notes,
                    }
                    rows.append(row)

                    if on_progress:
                        on_progress(done, total, row)

        finally:
            try:
                context.close()
            except Exception:
                pass

    out_full, out_missing = _write_outputs(rows, out_dir)
    return out_full, out_missing